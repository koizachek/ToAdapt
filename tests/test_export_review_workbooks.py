from __future__ import annotations

import importlib.util
import json
from pathlib import Path
import sys

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "export_review_workbooks.py"
SPEC = importlib.util.spec_from_file_location("export_review_workbooks", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC is not None and SPEC.loader is not None
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def test_export_review_workbooks_generates_rubric_and_blind_files(tmp_path):
    submissions = [
        {
            "submission_id": "sub-1",
            "user_id": "prolific_user_a",
            "matrikelnummer": "pid-a",
            "case_id": "case-alpha",
            "experiment": {
                "provider": "prolific",
                "prolific_pid": "pid-a",
                "prolific_session_id": "sess-a",
            },
            "answers": {
                "q1": "Antwort zu Frage 1",
                "q2": "Antwort zu Frage 2",
            },
            "scores": [
                {
                    "question_id": "q1",
                    "bloom_level": 4,
                    "max_points": 25,
                    "awarded_points": 19.5,
                    "feedback": "Gute Struktur",
                    "learning_objective_tags": ["analyse", "trade-off"],
                    "rubric_reference": "tp1_rubric.json",
                    "canvas_alignment_pct": 75.0,
                    "required_canvas_blocks": ["channels"],
                    "addressed_canvas_blocks": ["channels"],
                    "missing_canvas_blocks": [],
                    "canvas_rationale": "Passend angewendet.",
                    "evaluation_status": "ok",
                    "needs_human_review": False,
                    "review_reason": "",
                    "judge_confidence": "high",
                    "score_band": "strong",
                    "main_strengths": ["Case-Bezug"],
                    "main_penalties": [],
                },
                {
                    "question_id": "q2",
                    "bloom_level": 5,
                    "max_points": 24,
                    "awarded_points": 20.0,
                    "feedback": "Guter Trade-off",
                    "learning_objective_tags": ["evaluieren"],
                    "rubric_reference": "tp2_rubric.json",
                    "canvas_alignment_pct": 50.0,
                    "required_canvas_blocks": [],
                    "addressed_canvas_blocks": [],
                    "missing_canvas_blocks": [],
                    "canvas_rationale": "",
                    "evaluation_status": "technical_fallback",
                    "needs_human_review": True,
                    "review_reason": "Invalid JSON",
                    "judge_confidence": "low",
                    "score_band": "unscored",
                    "main_strengths": [],
                    "main_penalties": ["Technischer Fallback"],
                },
            ],
            "percentage": 82.4,
            "canvas_alignment_pct": 62.5,
            "rubric_fit_pct": 80.0,
            "started_at": "2026-05-17T09:00:00",
            "submitted_at": "2026-05-17T09:10:00",
            "evaluated_at": "2026-05-17T09:12:00",
        }
    ]
    submissions_path = tmp_path / "submission_states.json"
    submissions_path.write_text(json.dumps(submissions, ensure_ascii=False), encoding="utf-8")
    events = [
        {
            "event_type": "chat_turn_completed",
            "created_at": "2026-05-17 09:05:00",
            "payload": {
                "session_id": "sess-1",
                "case_id": "case-alpha",
                "user_id": "prolific_user_a",
                "experiment": {
                    "provider": "prolific",
                    "experiment_name": "pilot",
                    "run_id": "pid-a",
                    "prolific_pid": "pid-a",
                    "prolific_session_id": "sess-a",
                },
                "message_count": 1,
                "history_length": 1,
                "user_message": "Was ist der Kernkonflikt?",
                "agent_type": "metacognitive",
                "assistant_message": "Welche Spannung zwischen Kosten und Beratung siehst du?",
            },
        }
    ]
    events_path = tmp_path / "experiment_events.json"
    events_path.write_text(json.dumps(events, ensure_ascii=False), encoding="utf-8")

    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    case_payload = {
        "case_id": "case-alpha",
        "title": "Case Alpha",
        "questions": [
            {
                "question_id": "q1",
                "phase": 1,
                "bloom_level": 4,
                "text": "Frage eins?",
                "max_points": 25,
            },
            {
                "question_id": "q2",
                "phase": 2,
                "bloom_level": 5,
                "text": "Frage zwei?",
                "max_points": 24,
            },
        ],
    }
    (cases_dir / "case-alpha.json").write_text(json.dumps(case_payload, ensure_ascii=False), encoding="utf-8")

    output_dir = tmp_path / "out"
    paths = MODULE.export_review_workbooks(
        submissions_path=submissions_path,
        events_path=events_path,
        cases_dir=cases_dir,
        output_dir=output_dir,
        prefix="review",
    )

    rubric_wb = load_workbook(paths["rubric_workbook"])
    blind_wb = load_workbook(paths["blind_workbook"])
    chat_wb = load_workbook(paths["chat_turns_workbook"])

    assert rubric_wb.sheetnames == ["Overview", "P1_q1_case-alpha", "P2_q2_case-alpha"]
    assert blind_wb.sheetnames == ["Overview", "P1_q1_case-alpha", "P2_q2_case-alpha"]
    assert chat_wb.sheetnames == ["Overview", "ChatTurns"]

    rubric_sheet = rubric_wb["P1_q1_case-alpha"]
    blind_sheet = blind_wb["P1_q1_case-alpha"]

    assert rubric_sheet["A2"].value == "case-alpha:q1:001"
    assert rubric_sheet["C2"].value == "prolific_user_a"
    assert rubric_sheet["O2"].value == 19.5
    assert rubric_sheet["X2"].value == "ok"
    assert rubric_sheet["Y2"].value is False
    assert blind_sheet["A2"].value == "case-alpha:q1:001"
    assert blind_sheet["J2"].value is None
    assert blind_sheet["K2"].value is None
    chat_sheet = chat_wb["ChatTurns"]
    assert chat_sheet["A2"].value == "sess-1:001"
    assert chat_sheet["M2"].value == "metacognitive"
    assert "Kosten und Beratung" in chat_sheet["O2"].value
    assert all(header is not None for header in rubric_sheet[1])
