"""Tests für die Tutor-Antwort-Bewertung (NAACL-2025-Taxonomie, LLM-as-Judge)."""

import importlib.util
import json
from pathlib import Path

from backend.evaluator.tutor_eval import (
    TUTOR_EVAL_DIMENSIONS,
    TutorResponseEvaluator,
    aggregate_annotations,
    build_judge_prompt,
    normalize_annotation,
)

REPO_ROOT = Path(__file__).resolve().parents[1]


def _load_script():
    spec = importlib.util.spec_from_file_location(
        "evaluate_tutor_responses", REPO_ROOT / "scripts" / "evaluate_tutor_responses.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ---------------------------------------------------------------------------
# Taxonomie-Treue zum Framework
# ---------------------------------------------------------------------------

def test_taxonomy_has_eight_dimensions_with_paper_scales():
    assert list(TUTOR_EVAL_DIMENSIONS) == [
        "mistake_identification", "mistake_location", "revealing_of_the_answer",
        "providing_guidance", "actionability", "coherence", "tutor_tone",
        "humanlikeness",
    ]
    assert TUTOR_EVAL_DIMENSIONS["revealing_of_the_answer"]["labels"] == [
        "Yes (and correct)", "Yes (but incorrect)", "No",
    ]
    assert TUTOR_EVAL_DIMENSIONS["tutor_tone"]["labels"] == [
        "Encouraging", "Neutral", "Offensive",
    ]
    for dim in TUTOR_EVAL_DIMENSIONS.values():
        assert dim["desired"] in dim["labels"]
    # ToAdapt-Doktrin: Antwort-Enthüllung ist immer unerwünscht
    assert TUTOR_EVAL_DIMENSIONS["revealing_of_the_answer"]["desired"] == "No"


def test_judge_prompt_contains_all_dimensions_and_labels():
    prompt = build_judge_prompt(
        context="", student_message="Ich würde Option A nehmen.",
        tutor_response="Woran machst du fest, dass A besser trägt als B?",
    )
    for key, dim in TUTOR_EVAL_DIMENSIONS.items():
        assert key in prompt
        for label in dim["labels"]:
            assert label in prompt


# ---------------------------------------------------------------------------
# Normalisierung & Aggregation
# ---------------------------------------------------------------------------

def test_normalize_annotation_validates_labels():
    data = {
        "mistake_identification": {"label": "not applicable", "justification": "kein Fehler"},
        "revealing_of_the_answer": {"label": "NO", "justification": "nur Frage"},
        "tutor_tone": "Encouraging",              # String-Form erlaubt
        "coherence": {"label": "vielleicht"},      # ungültig
        # providing_guidance fehlt komplett
    }
    result = normalize_annotation(data)
    assert result["mistake_identification"]["label"] == "Not applicable"
    assert result["revealing_of_the_answer"]["label"] == "No"
    assert result["tutor_tone"]["label"] == "Encouraging"
    assert result["coherence"]["label"] == "Invalid"
    assert result["providing_guidance"]["label"] == "Invalid"


def _row(agent: str, **labels) -> dict:
    annotation = {
        key: {"label": labels.get(key, dim["desired"]), "justification": ""}
        for key, dim in TUTOR_EVAL_DIMENSIONS.items()
    }
    return {"agent_type": agent, "annotation": annotation}


def test_aggregate_excludes_na_and_invalid_from_desirability():
    rows = [
        _row("strategic"),                                             # alles desired
        _row("strategic", mistake_identification="Not applicable"),   # NA zählt nicht
        _row("conceptual", revealing_of_the_answer="Yes (and correct)",
             coherence="Invalid"),
    ]
    summary = aggregate_annotations(rows)

    mi = summary["overall"]["mistake_identification"]
    assert mi["n_valid"] == 2 and mi["desirability_rate"] == 1.0
    assert mi["counts"]["Not applicable"] == 1

    reveal = summary["overall"]["revealing_of_the_answer"]
    assert reveal["n_valid"] == 3
    assert reveal["desirability_rate"] == round(2 / 3, 3)

    coherence = summary["overall"]["coherence"]
    assert coherence["n_valid"] == 2  # Invalid ausgeschlossen

    assert set(summary["by_agent_type"]) == {"strategic", "conceptual"}
    assert summary["n_items"] == 3


async def test_evaluator_parses_judge_json(monkeypatch):
    reply = json.dumps({
        key: {"label": dim["desired"], "justification": "passt"}
        for key, dim in TUTOR_EVAL_DIMENSIONS.items()
    })

    async def fake_complete(self, *, system, messages, max_tokens):
        return reply

    monkeypatch.setattr("backend.llm.OpenRouterClient.complete", fake_complete)
    evaluator = TutorResponseEvaluator(api_key="test")
    annotation = await evaluator.evaluate_turn(
        context="", student_message="Frage?", tutor_response="Gegenfrage?",
    )
    assert annotation["providing_guidance"]["label"] == "Yes"
    assert annotation["tutor_tone"]["label"] == "Encouraging"


# ---------------------------------------------------------------------------
# Skript: Extraktion, Workbook, Aggregation
# ---------------------------------------------------------------------------

def _events() -> list[dict]:
    return [
        {"event_type": "chat_turn_completed", "payload": {
            "session_id": "sess-1", "case_id": "case-a", "agent_type": "strategic",
            "user_message": "Ich nehme Option A.", "assistant_message": "Warum A?",
        }},
        {"event_type": "chat_turn_completed", "payload": {
            "session_id": "sess-1", "case_id": "case-a", "agent_type": "conceptual",
            "user_message": "Was heißt Skalierung?", "assistant_message": "Was verbindest du damit?",
        }},
        # flaches Format (ohne payload-Hülle)
        {"event_type": "chat_turn_completed", "session_id": "sess-2",
         "agent_type": "metacognitive", "user_message": "Hilfe",
         "assistant_message": "Was ist euer Plan?"},
        {"event_type": "formative_feedback_requested", "payload": {
            "submission_id": "sub-1", "question_id": "q1", "request_number": 1,
            "draft_text": "Mein Entwurf.", "feedback": "Was folgt daraus?",
        }},
        {"event_type": "session_created", "payload": {"session_id": "sess-1"}},
    ]


def test_extract_tutor_items_builds_stable_ids():
    script = _load_script()
    items = script.extract_tutor_items(_events(), include_feedback=True)
    ids = [item["tutor_item_id"] for item in items]
    assert ids == ["sess-1:001", "sess-1:002", "sess-2:001", "sub-1:q1:1"]
    assert items[3]["agent_type"] == "formative_feedback"
    assert items[3]["student_message"] == "Mein Entwurf."

    without_feedback = script.extract_tutor_items(_events(), include_feedback=False)
    assert len(without_feedback) == 3


def test_annotation_workbook_written(tmp_path):
    from openpyxl import load_workbook

    script = _load_script()
    items = script.extract_tutor_items(_events())
    path = script.write_annotation_workbook(items, tmp_path / "blind.xlsx")

    workbook = load_workbook(path)
    sheet = workbook["annotation"]
    header = [cell.value for cell in sheet[1]]
    assert header[:5] == ["tutor_item_id", "source", "agent_type", "student_message", "tutor_response"]
    assert header[5:] == list(TUTOR_EVAL_DIMENSIONS)
    assert sheet.max_row == 1 + len(items)
    assert "labels" in workbook.sheetnames


def test_summary_written_from_jsonl(tmp_path):
    script = _load_script()
    rows = [_row("strategic"), _row("conceptual")]
    jsonl = tmp_path / "results.jsonl"
    script.write_results_jsonl(rows, jsonl)

    loaded = script.read_results_jsonl(jsonl)
    json_path, csv_path = script.write_summary(loaded, tmp_path, "20260709T000000Z")

    summary = json.loads(json_path.read_text(encoding="utf-8"))
    assert summary["n_items"] == 2
    csv_text = csv_path.read_text(encoding="utf-8")
    assert csv_text.startswith("scope;dimension;desired_label")
    assert "agent:strategic" in csv_text
