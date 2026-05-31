from __future__ import annotations

import importlib.util
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "compare_teacher_rubric_scores.py"
SPEC = importlib.util.spec_from_file_location("compare_teacher_rubric_scores", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC is not None and SPEC.loader is not None
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def _write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P1_q1_case-alpha"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_compare_teacher_rubric_scores_uses_teacher_rows_as_scope(tmp_path):
    teacher_path = tmp_path / "teacher.xlsx"
    rubric_path = tmp_path / "rubric.xlsx"

    _write_workbook(
        teacher_path,
        [
            "review_item_id",
            "question_id",
            "phase",
            "max_points",
            "answer_text",
            "teacher_awarded_points",
            "teacher_rationale",
        ],
        [
            ["case-alpha:q1:001", "q1", 1, 25, "Antwort A", 10, "Teilweise."],
            ["case-alpha:q1:002", "q1", 1, 25, "Antwort B", 20, "Stark."],
        ],
    )
    _write_workbook(
        rubric_path,
        [
            "review_item_id",
            "submission_id",
            "prolific_pid",
            "question_id",
            "max_points",
            "rubric_awarded_points",
            "rubric_feedback",
            "rubric_canvas_alignment_pct",
        ],
        [
            ["case-alpha:q1:001", "sub-1", "pid-1", "q1", 25, 12, "ok", 50],
            ["case-alpha:q1:002", "sub-2", "pid-2", "q1", 25, 18, "ok", 80],
            ["case-alpha:q1:003", "sub-test", "pid-test", "q1", 25, 25, "excluded", 100],
        ],
    )

    paths = MODULE.compare_teacher_rubric_scores(
        teacher_workbook=teacher_path,
        rubric_workbook=rubric_path,
        output_dir=tmp_path,
        prefix="comparison",
    )

    workbook = load_workbook(paths.workbook, data_only=True)
    summary = {
        row[0]: row[1]
        for row in workbook["summary"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    validation_rows = list(workbook["validation"].iter_rows(min_row=2, values_only=True))

    assert summary["matched_scored_rows_used_for_metrics"] == 2
    assert summary["rubric_only_ids_excluded_by_design"] == 1
    assert ("rubric_only_excluded", "case-alpha:q1:003", None) in validation_rows
