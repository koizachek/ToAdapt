"""Erzeugt zwei Excel-Workbooks fuer Rubric- und Blind-Reviews."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SUBMISSIONS_PATH = REPO_ROOT / "data" / "submission_states.json"
DEFAULT_EVENTS_PATH = REPO_ROOT / "data" / "experiment_events.json"
DEFAULT_CASES_DIR = REPO_ROOT / "backend" / "cases" / "pool"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "prolific_runs" / "derived" / "review_exports"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
TOP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class QuestionMeta:
    case_id: str
    case_title: str
    question_id: str
    phase: int
    bloom_level: int
    max_points: int
    question_text: str


@dataclass(frozen=True)
class ReviewRow:
    review_item_id: str
    submission_id: str
    user_id: str
    participant_id: str
    prolific_pid: str
    prolific_session_id: str
    case_id: str
    case_title: str
    question_id: str
    phase: int
    bloom_level: int
    max_points: int
    question_text: str
    answer_text: str
    rubric_awarded_points: float | None
    rubric_feedback: str
    rubric_learning_objective_tags: str
    rubric_reference: str
    rubric_canvas_alignment_pct: float | None
    rubric_required_canvas_blocks: str
    rubric_addressed_canvas_blocks: str
    rubric_missing_canvas_blocks: str
    rubric_canvas_rationale: str
    submission_percentage: float | None
    submission_canvas_alignment_pct: float | None
    submission_rubric_fit_pct: float | None
    started_at: str
    submitted_at: str
    evaluated_at: str


@dataclass(frozen=True)
class ChatTurnRow:
    turn_id: str
    created_at: str
    session_id: str
    case_id: str
    user_id: str
    participant_id: str
    prolific_pid: str
    prolific_session_id: str
    experiment_name: str
    run_id: str
    message_count: int | None
    history_length: int | None
    agent_type: str
    user_message: str
    assistant_message: str


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_case_map(cases_dir: Path) -> dict[tuple[str, str], QuestionMeta]:
    mapping: dict[tuple[str, str], QuestionMeta] = {}
    for path in sorted(cases_dir.glob("*.json")):
        payload = _load_json(path)
        case_id = payload.get("case_id")
        case_title = payload.get("title")
        questions = payload.get("questions")
        if not case_id or not case_title or not isinstance(questions, list):
            continue
        for question in payload.get("questions", []):
            meta = QuestionMeta(
                case_id=case_id,
                case_title=case_title,
                question_id=question["question_id"],
                phase=question["phase"],
                bloom_level=question["bloom_level"],
                max_points=question["max_points"],
                question_text=question["text"],
            )
            mapping[(case_id, meta.question_id)] = meta
    return mapping


def _safe_sheet_name(base: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[:\\\\/?*\\[\\]]", "-", base).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate in used_names:
        tail = f"_{suffix}"
        candidate = f"{cleaned[:31 - len(tail)]}{tail}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _fmt_timestamp(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value)


def _join_list(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return str(value)


def _build_rows(submissions_path: Path, cases_dir: Path) -> list[ReviewRow]:
    submissions = _load_json(submissions_path)
    case_map = _load_case_map(cases_dir)
    rows: list[ReviewRow] = []
    sheet_counters: dict[tuple[str, str], int] = {}

    for submission in submissions:
        answers = submission.get("answers") or {}
        if not answers:
            continue

        score_by_question = {
            score["question_id"]: score
            for score in submission.get("scores", [])
            if score.get("question_id")
        }
        experiment = submission.get("experiment") or {}
        case_id = submission.get("case_id", "")

        for question_id, answer_text in answers.items():
            if not str(answer_text).strip():
                continue

            meta = case_map.get((case_id, question_id))
            score = score_by_question.get(question_id, {})
            key = (case_id, question_id)
            sheet_counters[key] = sheet_counters.get(key, 0) + 1
            review_item_id = f"{case_id}:{question_id}:{sheet_counters[key]:03d}"

            rows.append(
                ReviewRow(
                    review_item_id=review_item_id,
                    submission_id=submission.get("submission_id", ""),
                    user_id=submission.get("user_id", ""),
                    participant_id=submission.get("matrikelnummer", ""),
                    prolific_pid=experiment.get("prolific_pid", ""),
                    prolific_session_id=experiment.get("prolific_session_id", ""),
                    case_id=case_id,
                    case_title=(meta.case_title if meta else case_id),
                    question_id=question_id,
                    phase=(meta.phase if meta else score.get("phase") or 0),
                    bloom_level=(meta.bloom_level if meta else score.get("bloom_level") or 0),
                    max_points=(meta.max_points if meta else score.get("max_points") or 0),
                    question_text=(meta.question_text if meta else ""),
                    answer_text=str(answer_text),
                    rubric_awarded_points=score.get("awarded_points"),
                    rubric_feedback=score.get("feedback", ""),
                    rubric_learning_objective_tags=_join_list(score.get("learning_objective_tags")),
                    rubric_reference=score.get("rubric_reference", ""),
                    rubric_canvas_alignment_pct=score.get("canvas_alignment_pct"),
                    rubric_required_canvas_blocks=_join_list(score.get("required_canvas_blocks")),
                    rubric_addressed_canvas_blocks=_join_list(score.get("addressed_canvas_blocks")),
                    rubric_missing_canvas_blocks=_join_list(score.get("missing_canvas_blocks")),
                    rubric_canvas_rationale=score.get("canvas_rationale", ""),
                    submission_percentage=submission.get("percentage"),
                    submission_canvas_alignment_pct=submission.get("canvas_alignment_pct"),
                    submission_rubric_fit_pct=submission.get("rubric_fit_pct"),
                    started_at=_fmt_timestamp(submission.get("started_at")),
                    submitted_at=_fmt_timestamp(submission.get("submitted_at")),
                    evaluated_at=_fmt_timestamp(submission.get("evaluated_at")),
                )
            )

    rows.sort(key=lambda row: (row.case_id, row.phase, row.question_id, row.review_item_id))
    return rows


def _group_rows(rows: list[ReviewRow]) -> dict[tuple[str, str], list[ReviewRow]]:
    grouped: dict[tuple[str, str], list[ReviewRow]] = {}
    for row in rows:
        grouped.setdefault((row.case_id, row.question_id), []).append(row)
    return grouped


def _build_chat_turn_rows(events_path: Path) -> list[ChatTurnRow]:
    events = _load_json(events_path)
    if not isinstance(events, list):
        raise ValueError("Event-Datei muss eine JSON-Liste enthalten.")

    rows: list[ChatTurnRow] = []
    for index, event in enumerate(events, start=1):
        if event.get("event_type") != "chat_turn_completed":
            continue

        payload = event.get("payload") or {}
        experiment = payload.get("experiment") or {}
        session_id = payload.get("session_id", "")
        message_count = payload.get("message_count")
        turn_id = (
            f"{session_id}:{int(message_count):03d}"
            if session_id and isinstance(message_count, int)
            else f"chat_turn:{index:04d}"
        )

        rows.append(
            ChatTurnRow(
                turn_id=turn_id,
                created_at=_fmt_timestamp(event.get("created_at")),
                session_id=session_id,
                case_id=payload.get("case_id", ""),
                user_id=payload.get("user_id", ""),
                participant_id=experiment.get("prolific_pid", ""),
                prolific_pid=experiment.get("prolific_pid", ""),
                prolific_session_id=experiment.get("prolific_session_id", ""),
                experiment_name=experiment.get("experiment_name", ""),
                run_id=experiment.get("run_id", ""),
                message_count=message_count if isinstance(message_count, int) else None,
                history_length=payload.get("history_length")
                if isinstance(payload.get("history_length"), int)
                else None,
                agent_type=payload.get("agent_type", ""),
                user_message=payload.get("user_message", ""),
                assistant_message=payload.get("assistant_message", ""),
            )
        )

    rows.sort(
        key=lambda row: (
            row.participant_id or row.user_id,
            row.session_id,
            row.message_count if row.message_count is not None else 10**9,
            row.created_at,
        )
    )
    return rows


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGNMENT

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP_ALIGNMENT


def _apply_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _write_rubric_workbook(rows: list[ReviewRow], output_path: Path) -> None:
    workbook = Workbook()
    _add_overview_sheet(rows, [], workbook, title="Overview", blind=False, chat_turns=False)
    used_names = {"Overview"}
    grouped = _group_rows(rows)

    headers = [
        "review_item_id",
        "submission_id",
        "user_id",
        "participant_id",
        "prolific_pid",
        "prolific_session_id",
        "case_id",
        "case_title",
        "question_id",
        "phase",
        "bloom_level",
        "max_points",
        "question_text",
        "answer_text",
        "rubric_awarded_points",
        "rubric_feedback",
        "rubric_learning_objective_tags",
        "rubric_reference",
        "rubric_canvas_alignment_pct",
        "rubric_required_canvas_blocks",
        "rubric_addressed_canvas_blocks",
        "rubric_missing_canvas_blocks",
        "rubric_canvas_rationale",
        "submission_percentage",
        "submission_canvas_alignment_pct",
        "submission_rubric_fit_pct",
        "started_at",
        "submitted_at",
        "evaluated_at",
    ]

    for (_, _), sheet_rows in grouped.items():
        first = sheet_rows[0]
        title = _safe_sheet_name(f"P{first.phase}_{first.question_id}_{first.case_id}", used_names)
        ws = workbook.create_sheet(title)
        ws.append(headers)
        for row in sheet_rows:
            ws.append([
                row.review_item_id,
                row.submission_id,
                row.user_id,
                row.participant_id,
                row.prolific_pid,
                row.prolific_session_id,
                row.case_id,
                row.case_title,
                row.question_id,
                row.phase,
                row.bloom_level,
                row.max_points,
                row.question_text,
                row.answer_text,
                row.rubric_awarded_points,
                row.rubric_feedback,
                row.rubric_learning_objective_tags,
                row.rubric_reference,
                row.rubric_canvas_alignment_pct,
                row.rubric_required_canvas_blocks,
                row.rubric_addressed_canvas_blocks,
                row.rubric_missing_canvas_blocks,
                row.rubric_canvas_rationale,
                row.submission_percentage,
                row.submission_canvas_alignment_pct,
                row.submission_rubric_fit_pct,
                row.started_at,
                row.submitted_at,
                row.evaluated_at,
            ])
        _style_sheet(ws)
        _apply_widths(
            ws,
            {
                "A": 24,
                "B": 38,
                "C": 20,
                "D": 18,
                "E": 18,
                "F": 22,
                "G": 24,
                "H": 34,
                "I": 10,
                "J": 8,
                "K": 12,
                "L": 12,
                "M": 60,
                "N": 90,
                "O": 18,
                "P": 70,
                "Q": 28,
                "R": 20,
                "S": 22,
                "T": 28,
                "U": 28,
                "V": 28,
                "W": 45,
                "X": 16,
                "Y": 24,
                "Z": 22,
                "AA": 22,
                "AB": 22,
                "AC": 22,
            },
        )

    workbook.save(output_path)


def _write_blind_workbook(rows: list[ReviewRow], output_path: Path) -> None:
    workbook = Workbook()
    _add_overview_sheet(rows, [], workbook, title="Overview", blind=True, chat_turns=False)
    used_names = {"Overview"}
    grouped = _group_rows(rows)

    headers = [
        "review_item_id",
        "case_id",
        "case_title",
        "question_id",
        "phase",
        "bloom_level",
        "max_points",
        "question_text",
        "answer_text",
        "teacher_awarded_points",
        "teacher_rationale",
    ]

    for (_, _), sheet_rows in grouped.items():
        first = sheet_rows[0]
        title = _safe_sheet_name(f"P{first.phase}_{first.question_id}_{first.case_id}", used_names)
        ws = workbook.create_sheet(title)
        ws.append(headers)
        for row in sheet_rows:
            ws.append([
                row.review_item_id,
                row.case_id,
                row.case_title,
                row.question_id,
                row.phase,
                row.bloom_level,
                row.max_points,
                row.question_text,
                row.answer_text,
                "",
                "",
            ])
        _style_sheet(ws)
        _apply_widths(
            ws,
            {
                "A": 24,
                "B": 24,
                "C": 34,
                "D": 10,
                "E": 8,
                "F": 12,
                "G": 12,
                "H": 60,
                "I": 90,
                "J": 18,
                "K": 70,
            },
        )

    workbook.save(output_path)


def _write_chat_turn_workbook(rows: list[ChatTurnRow], output_path: Path) -> None:
    workbook = Workbook()
    _add_overview_sheet([], rows, workbook, title="Overview", blind=False, chat_turns=True)

    ws = workbook.create_sheet("ChatTurns")
    headers = [
        "turn_id",
        "created_at",
        "session_id",
        "case_id",
        "user_id",
        "participant_id",
        "prolific_pid",
        "prolific_session_id",
        "experiment_name",
        "run_id",
        "message_count",
        "history_length",
        "agent_type",
        "user_message",
        "assistant_message",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.turn_id,
            row.created_at,
            row.session_id,
            row.case_id,
            row.user_id,
            row.participant_id,
            row.prolific_pid,
            row.prolific_session_id,
            row.experiment_name,
            row.run_id,
            row.message_count,
            row.history_length,
            row.agent_type,
            row.user_message,
            row.assistant_message,
        ])
    _style_sheet(ws)
    _apply_widths(
        ws,
        {
            "A": 30,
            "B": 24,
            "C": 38,
            "D": 24,
            "E": 22,
            "F": 18,
            "G": 18,
            "H": 22,
            "I": 28,
            "J": 18,
            "K": 14,
            "L": 14,
            "M": 16,
            "N": 70,
            "O": 90,
        },
    )

    workbook.save(output_path)


def _add_overview_sheet(
    review_rows: list[ReviewRow],
    chat_rows: list[ChatTurnRow],
    workbook: Workbook,
    *,
    title: str,
    blind: bool,
    chat_turns: bool,
) -> None:
    ws = workbook.active
    ws.title = title
    ws.append(["exported_at_utc", datetime.now(timezone.utc).isoformat()])
    if chat_turns:
        ws.append(["row_count", len(chat_rows)])
        ws.append(["sheet_strategy", "ein Blatt mit einer Zeile pro Chat-Turn"])
        ws.append(["contains_user_messages", "ja"])
        ws.append(["contains_assistant_messages", "ja"])
        ws.append(
            [
                "hinweis",
                "Chat-Turn-Datei basiert auf experiment_events.json und enthaelt nur "
                "`chat_turn_completed`-Events.",
            ]
        )
    else:
        ws.append(["row_count", len(review_rows)])
        ws.append(["sheet_strategy", "ein Blatt pro Frage"])
        ws.append(["blind_mode", "ja" if blind else "nein"])
        if blind:
            ws.append(
                [
                    "hinweis",
                    "Blind-Datei enthaelt keine user_id, participant_id oder Rubric-Bewertungen. "
                    "Die Zuordnung fuer spaeteren Abgleich laeuft ueber review_item_id.",
                ]
            )
        else:
            ws.append(
                [
                    "hinweis",
                    "Rubric-Datei enthaelt dieselbe review_item_id wie die Blind-Datei und kann "
                    "spaeter mit den menschlichen Bewertungen gemerged werden.",
                ]
            )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = TOP_ALIGNMENT
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 120


def export_review_workbooks(
    *,
    submissions_path: Path,
    events_path: Path | None,
    cases_dir: Path,
    output_dir: Path,
    prefix: str,
) -> dict[str, Path]:
    rows = _build_rows(submissions_path, cases_dir)
    if not rows:
        raise ValueError("Keine Antworten in den Submission-Daten gefunden.")

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    rubric_path = output_dir / f"{prefix}_{timestamp}_rubric.xlsx"
    blind_path = output_dir / f"{prefix}_{timestamp}_blind.xlsx"
    chat_turns_path = output_dir / f"{prefix}_{timestamp}_chat_turns.xlsx"

    _write_rubric_workbook(rows, rubric_path)
    _write_blind_workbook(rows, blind_path)
    paths = {
        "rubric_workbook": rubric_path,
        "blind_workbook": blind_path,
    }

    if events_path and events_path.exists():
        chat_rows = _build_chat_turn_rows(events_path)
        if chat_rows:
            _write_chat_turn_workbook(chat_rows, chat_turns_path)
            paths["chat_turns_workbook"] = chat_turns_path

    return paths


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--submissions",
        default=str(DEFAULT_SUBMISSIONS_PATH),
        help="Pfad zur JSON-Datei mit Submission-Zustaenden",
    )
    parser.add_argument(
        "--cases-dir",
        default=str(DEFAULT_CASES_DIR),
        help="Verzeichnis mit den Case-JSON-Dateien",
    )
    parser.add_argument(
        "--events",
        default=str(DEFAULT_EVENTS_PATH),
        help="Pfad zur JSON-Datei mit Experiment-Events",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Zielverzeichnis fuer die generierten Excel-Dateien",
    )
    parser.add_argument(
        "--prefix",
        default="prolific_review",
        help="Dateipraefix fuer die Excel-Exporte",
    )
    args = parser.parse_args()

    paths = export_review_workbooks(
        submissions_path=Path(args.submissions).expanduser().resolve(),
        events_path=Path(args.events).expanduser().resolve(),
        cases_dir=Path(args.cases_dir).expanduser().resolve(),
        output_dir=Path(args.output_dir).expanduser().resolve(),
        prefix=args.prefix,
    )
    print(paths["rubric_workbook"])
    print(paths["blind_workbook"])
    if "chat_turns_workbook" in paths:
        print(paths["chat_turns_workbook"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
