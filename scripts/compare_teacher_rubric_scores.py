"""Vergleicht bereinigte Lehrerbewertungen mit Rubric-Scores.

Die Lehrerdatei ist die kanonische Stichprobe. Rubric-Zeilen ohne passende
review_item_id werden bewusst ausgeschlossen, damit bereinigte Testuser nicht
wieder in die Analyse geraten.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime, timezone
from math import sqrt
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "prolific_runs" / "derived"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
TOP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class ComparisonPaths:
    workbook: Path
    csv: Path


def _normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() in {"overview", "task"}:
            continue

        sheet = workbook[sheet_name]
        headers = [_normalize_header(cell.value) for cell in sheet[1]]
        if "review_item_id" not in headers:
            continue

        for row_index, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row = dict(zip(headers, values))
            if not row.get("review_item_id"):
                continue
            row["_sheet"] = sheet_name
            row["_row"] = row_index
            rows.append(row)
    return rows


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def _pearson(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 2:
        return None
    x_mean = mean(xs)
    y_mean = mean(ys)
    numerator = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys))
    x_denominator = sqrt(sum((x - x_mean) ** 2 for x in xs))
    y_denominator = sqrt(sum((y - y_mean) ** 2 for y in ys))
    if x_denominator == 0 or y_denominator == 0:
        return None
    return numerator / (x_denominator * y_denominator)


def _round(value: float | None, digits: int = 3) -> float | None:
    return None if value is None else round(value, digits)


def _write_sheet(
    sheet,
    headers: list[str],
    rows: list[dict[str, Any]],
    widths: dict[str, float] | None = None,
) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGNMENT

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP_ALIGNMENT

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in (widths or {}).items():
        sheet.column_dimensions[column].width = width


def _csv_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if any(char in text for char in [";", '"', "\n"]):
        return '"' + text.replace('"', '""') + '"'
    return text


def _metrics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {
            "n": 0,
            "teacher_mean": None,
            "rubric_mean": None,
            "mean_diff_rubric_minus_teacher": None,
            "median_diff": None,
            "mae_points": None,
            "rmse_points": None,
            "mean_abs_diff_pct_points": None,
            "pearson_r": None,
            "rubric_higher_n": 0,
            "teacher_higher_n": 0,
            "equal_n": 0,
            "within_1pt_n": 0,
            "within_2pt_n": 0,
            "within_10pctpts_n": 0,
        }

    teacher = [row["teacher_awarded_points"] for row in rows]
    rubric = [row["rubric_awarded_points"] for row in rows]
    diffs = [row["diff_points_rubric_minus_teacher"] for row in rows]
    abs_diffs = [row["abs_diff_points"] for row in rows]
    pct_abs_diffs = [
        abs(row["diff_pct_points"])
        for row in rows
        if row["diff_pct_points"] is not None
    ]
    return {
        "n": len(rows),
        "teacher_mean": round(mean(teacher), 3),
        "rubric_mean": round(mean(rubric), 3),
        "mean_diff_rubric_minus_teacher": round(mean(diffs), 3),
        "median_diff": round(median(diffs), 3),
        "mae_points": round(mean(abs_diffs), 3),
        "rmse_points": round(sqrt(mean([diff**2 for diff in diffs])), 3),
        "mean_abs_diff_pct_points": round(mean(pct_abs_diffs), 3)
        if pct_abs_diffs
        else None,
        "pearson_r": _round(_pearson(teacher, rubric), 3),
        "rubric_higher_n": sum(1 for row in rows if row["direction"] == "rubric_higher"),
        "teacher_higher_n": sum(1 for row in rows if row["direction"] == "teacher_higher"),
        "equal_n": sum(1 for row in rows if row["direction"] == "equal"),
        "within_1pt_n": sum(1 for row in rows if row["abs_diff_points"] <= 1),
        "within_2pt_n": sum(1 for row in rows if row["abs_diff_points"] <= 2),
        "within_10pctpts_n": sum(1 for row in rows if abs(row["diff_pct_points"]) <= 10),
    }


def _index_rows(
    rows: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], list[str]]:
    indexed: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for row in rows:
        review_item_id = str(row["review_item_id"])
        if review_item_id in indexed:
            duplicates.append(review_item_id)
        indexed[review_item_id] = row
    return indexed, duplicates


def _build_comparison_rows(
    teacher_by_id: dict[str, dict[str, Any]],
    rubric_by_id: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    comparison_rows: list[dict[str, Any]] = []
    for review_item_id, teacher in teacher_by_id.items():
        rubric = rubric_by_id.get(review_item_id)
        teacher_points = _as_float(teacher.get("teacher_awarded_points"))
        rubric_points = _as_float(rubric.get("rubric_awarded_points")) if rubric else None
        max_points = _as_float(teacher.get("max_points"))
        teacher_pct = (
            teacher_points / max_points * 100
            if teacher_points is not None and max_points
            else None
        )
        rubric_pct = (
            rubric_points / max_points * 100
            if rubric_points is not None and max_points
            else None
        )
        diff = (
            rubric_points - teacher_points
            if teacher_points is not None and rubric_points is not None
            else None
        )
        diff_pct = (
            rubric_pct - teacher_pct
            if teacher_pct is not None and rubric_pct is not None
            else None
        )
        abs_diff = abs(diff) if diff is not None else None

        direction = ""
        if diff is not None:
            if diff > 0:
                direction = "rubric_higher"
            elif diff < 0:
                direction = "teacher_higher"
            else:
                direction = "equal"

        issue_flags: list[str] = []
        if not rubric:
            issue_flags.append("missing_rubric_match")
        if teacher_points is None:
            issue_flags.append("missing_teacher_score")
        if rubric_points is None:
            issue_flags.append("missing_rubric_score")
        if teacher_points is not None and max_points is not None:
            if not 0 <= teacher_points <= max_points:
                issue_flags.append("teacher_score_out_of_range")
        if rubric_points is not None and max_points is not None:
            if not 0 <= rubric_points <= max_points:
                issue_flags.append("rubric_score_out_of_range")
        if rubric and teacher.get("question_id") != rubric.get("question_id"):
            issue_flags.append("question_id_mismatch")
        if rubric and _as_float(teacher.get("max_points")) != _as_float(rubric.get("max_points")):
            issue_flags.append("max_points_mismatch")

        comparison_rows.append(
            {
                "review_item_id": review_item_id,
                "question_id": teacher.get("question_id"),
                "phase": teacher.get("phase"),
                "max_points": max_points,
                "teacher_awarded_points": teacher_points,
                "rubric_awarded_points": rubric_points,
                "diff_points_rubric_minus_teacher": diff,
                "abs_diff_points": abs_diff,
                "teacher_pct_of_question": _round(teacher_pct, 2),
                "rubric_pct_of_question": _round(rubric_pct, 2),
                "diff_pct_points": _round(diff_pct, 2),
                "direction": direction,
                "rubric_canvas_alignment_pct": rubric.get("rubric_canvas_alignment_pct")
                if rubric
                else "",
                "rubric_evaluation_status": rubric.get("rubric_evaluation_status")
                if rubric
                else "",
                "rubric_needs_human_review": rubric.get("rubric_needs_human_review")
                if rubric
                else "",
                "rubric_review_reason": rubric.get("rubric_review_reason") if rubric else "",
                "rubric_judge_confidence": rubric.get("rubric_judge_confidence")
                if rubric
                else "",
                "rubric_score_band": rubric.get("rubric_score_band") if rubric else "",
                "rubric_required_canvas_blocks": rubric.get("rubric_required_canvas_blocks")
                if rubric
                else "",
                "rubric_addressed_canvas_blocks": rubric.get("rubric_addressed_canvas_blocks")
                if rubric
                else "",
                "rubric_missing_canvas_blocks": rubric.get("rubric_missing_canvas_blocks")
                if rubric
                else "",
                "teacher_rationale": teacher.get("teacher_rationale") or "",
                "rubric_feedback": rubric.get("rubric_feedback") if rubric else "",
                "rubric_canvas_rationale": rubric.get("rubric_canvas_rationale") if rubric else "",
                "answer_text": teacher.get("answer_text") or "",
                "submission_id": rubric.get("submission_id") if rubric else "",
                "prolific_pid": rubric.get("prolific_pid") if rubric else "",
                "issue_flags": "; ".join(issue_flags),
                "teacher_sheet": teacher.get("_sheet"),
                "teacher_row": teacher.get("_row"),
            }
        )
    return comparison_rows


def compare_teacher_rubric_scores(
    *,
    teacher_workbook: Path,
    rubric_workbook: Path,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    prefix: str = "teacher_rubric_comparison",
) -> ComparisonPaths:
    teacher_rows, rubric_rows = (
        _rows_from_workbook(teacher_workbook),
        _rows_from_workbook(rubric_workbook),
    )
    teacher_by_id, duplicate_teacher_ids = _index_rows(teacher_rows)
    rubric_by_id, duplicate_rubric_ids = _index_rows(rubric_rows)
    comparison_rows = _build_comparison_rows(teacher_by_id, rubric_by_id)

    scored_rows = [
        row
        for row in comparison_rows
        if row["teacher_awarded_points"] is not None
        and row["rubric_awarded_points"] is not None
    ]
    rubric_only_ids = sorted(set(rubric_by_id) - set(teacher_by_id))
    teacher_only_ids = sorted(set(teacher_by_id) - set(rubric_by_id))

    summary_rows = [
        {"metric": "teacher_file", "value": str(teacher_workbook)},
        {"metric": "rubric_file", "value": str(rubric_workbook)},
        {"metric": "exported_at_utc", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "teacher_rows_raw", "value": len(teacher_rows)},
        {"metric": "teacher_unique_review_item_ids", "value": len(teacher_by_id)},
        {
            "metric": "teacher_rows_with_scores",
            "value": sum(
                1 for row in comparison_rows if row["teacher_awarded_points"] is not None
            ),
        },
        {"metric": "rubric_rows_raw", "value": len(rubric_rows)},
        {"metric": "matched_scored_rows_used_for_metrics", "value": len(scored_rows)},
        {"metric": "teacher_only_ids", "value": len(teacher_only_ids)},
        {"metric": "rubric_only_ids_excluded_by_design", "value": len(rubric_only_ids)},
        {"metric": "duplicate_teacher_ids", "value": len(duplicate_teacher_ids)},
        {"metric": "duplicate_rubric_ids", "value": len(duplicate_rubric_ids)},
    ]
    summary_rows.extend(
        {"metric": key, "value": value} for key, value in _metrics(scored_rows).items()
    )
    summary_rows.append(
        {
            "metric": "scope_note",
            "value": (
                "Teacher workbook is canonical; rubric-only rows were excluded so "
                "cleaned/test users are not reintroduced."
            ),
        }
    )

    by_question_rows = []
    for question_id in sorted({row["question_id"] for row in scored_rows}):
        question_rows = [row for row in scored_rows if row["question_id"] == question_id]
        by_question_rows.append({"question_id": question_id, **_metrics(question_rows)})

    outlier_rows = sorted(
        scored_rows,
        key=lambda row: row["abs_diff_points"] if row["abs_diff_points"] is not None else -1,
        reverse=True,
    )[:20]

    validation_rows: list[dict[str, Any]] = []
    validation_rows.extend(
        {"type": "teacher_only_no_rubric_match", "review_item_id": item_id}
        for item_id in teacher_only_ids
    )
    validation_rows.extend(
        {"type": "rubric_only_excluded", "review_item_id": item_id}
        for item_id in rubric_only_ids
    )
    validation_rows.extend(
        {"type": "duplicate_teacher_id", "review_item_id": item_id}
        for item_id in duplicate_teacher_ids
    )
    validation_rows.extend(
        {"type": "duplicate_rubric_id", "review_item_id": item_id}
        for item_id in duplicate_rubric_ids
    )
    validation_rows.extend(
        {
            "type": "row_issue",
            "review_item_id": row["review_item_id"],
            "details": row["issue_flags"],
        }
        for row in comparison_rows
        if row["issue_flags"]
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / f"{prefix}.xlsx"
    csv_path = output_dir / f"{prefix}_item_comparison.csv"

    item_headers = [
        "review_item_id",
        "question_id",
        "phase",
        "max_points",
        "teacher_awarded_points",
        "rubric_awarded_points",
        "diff_points_rubric_minus_teacher",
        "abs_diff_points",
        "teacher_pct_of_question",
        "rubric_pct_of_question",
        "diff_pct_points",
        "direction",
        "rubric_canvas_alignment_pct",
        "rubric_evaluation_status",
        "rubric_needs_human_review",
        "rubric_review_reason",
        "rubric_judge_confidence",
        "rubric_score_band",
        "rubric_required_canvas_blocks",
        "rubric_addressed_canvas_blocks",
        "rubric_missing_canvas_blocks",
        "teacher_rationale",
        "rubric_feedback",
        "rubric_canvas_rationale",
        "answer_text",
        "submission_id",
        "prolific_pid",
        "issue_flags",
        "teacher_sheet",
        "teacher_row",
    ]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _write_sheet(summary_sheet, ["metric", "value"], summary_rows, {"A": 38, "B": 120})

    question_sheet = workbook.create_sheet("by_question")
    _write_sheet(
        question_sheet,
        [
            "question_id",
            "n",
            "teacher_mean",
            "rubric_mean",
            "mean_diff_rubric_minus_teacher",
            "median_diff",
            "mae_points",
            "rmse_points",
            "mean_abs_diff_pct_points",
            "pearson_r",
            "rubric_higher_n",
            "teacher_higher_n",
            "equal_n",
            "within_1pt_n",
            "within_2pt_n",
            "within_10pctpts_n",
        ],
        by_question_rows,
    )

    item_sheet = workbook.create_sheet("item_comparison")
    _write_sheet(item_sheet, item_headers, comparison_rows)

    outlier_sheet = workbook.create_sheet("outliers")
    _write_sheet(outlier_sheet, item_headers, outlier_rows)

    validation_sheet = workbook.create_sheet("validation")
    _write_sheet(validation_sheet, ["type", "review_item_id", "details"], validation_rows)
    workbook.save(workbook_path)

    csv_lines = [";".join(item_headers)]
    csv_lines.extend(
        ";".join(_csv_escape(row.get(header)) for header in item_headers)
        for row in comparison_rows
    )
    csv_path.write_text("\n".join(csv_lines) + "\n", encoding="utf-8")

    return ComparisonPaths(workbook=workbook_path, csv=csv_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--teacher-workbook", required=True)
    parser.add_argument("--rubric-workbook", required=True)
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Zielverzeichnis fuer Vergleichsdateien",
    )
    parser.add_argument(
        "--prefix",
        default="teacher_rubric_comparison",
        help="Dateipraefix ohne Dateiendung",
    )
    args = parser.parse_args()

    paths = compare_teacher_rubric_scores(
        teacher_workbook=Path(args.teacher_workbook).expanduser().resolve(),
        rubric_workbook=Path(args.rubric_workbook).expanduser().resolve(),
        output_dir=Path(args.output_dir).expanduser().resolve(),
        prefix=args.prefix,
    )
    print(paths.workbook)
    print(paths.csv)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
